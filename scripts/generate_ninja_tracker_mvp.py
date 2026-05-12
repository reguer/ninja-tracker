from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).resolve().parents[1]
OUT = BASE_DIR / 'Ninja_Tracker_MVP.xlsx'
HF = PatternFill('solid', fgColor='1F4E78')
HFF = Font(color='FFFFFF', bold=True)
SF = PatternFill('solid', fgColor='D9E1F2')
B = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

def hdr(ws, r, vals):
    for c, v in enumerate(vals, 1):
        x = ws.cell(r, c, v)
        x.fill, x.font, x.border = HF, HFF, B
        x.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def widths(ws, d):
    for k, v in d.items(): ws.column_dimensions[k].width = v

def grid(ws, r1, r2, c2):
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=1, max_col=c2):
        for cell in row: cell.border = B

def ps(col):
    return f'=IF($B$3="semanal",SUMIFS(Calendario!${col}:${col},Calendario!$D:$D,ISOWEEKNUM($B$4),Calendario!$F:$F,YEAR($B$4)),IF($B$3="mensual",SUMIFS(Calendario!${col}:${col},Calendario!$E:$E,MONTH($B$4),Calendario!$F:$F,YEAR($B$4)),SUMIFS(Calendario!${col}:${col},Calendario!$F:$F,YEAR($B$4))))'

def cfg(ws):
    ws.title = 'Config'; hdr(ws,1,['Parametro','Valor','Descripcion'])
    rows=[('FechaInicioCalendario','2026-01-01','Fecha base'),('HorizonteDias',365,'Dias proyectados'),('HorasBaseDia',24,'Horas teoricas/dia'),('HorasSuenoObjetivo',8,'Sueno diario'),('HorasFijasBaseInevitables',1.5,'Base fija diaria'),('DiasLaborables','1,2,3,4,5','Lunes=1...Domingo=7'),('UnidadBase','horas','Unidad base'),('VistaDefault','semanal','semanal/mensual/anual'),('InicioSemana','lunes','Referencia UX'),('BufferSeguridadDia',1,'Umbral riesgo'),('EscenarioSemanaActivo','=Dashboard!$B$5','Escenario activo')]
    for i,(a,b,c) in enumerate(rows,2): ws.cell(i,1,a); ws.cell(i,2,b); ws.cell(i,3,c)
    ws['B2'].number_format='yyyy-mm-dd'; ws.freeze_panes='A2'; widths(ws,{'A':34,'B':22,'C':64}); grid(ws,2,1+len(rows),3)

def cat(wb):
    ws=wb.create_sheet('Catálogos'); hdr(ws,1,['TiposActividad','TiposBloque','EsFijo','ConsumeTiempo','Frecuencias','Prioridades','PrioridadReal','EstadosProyecto','EstadosFase','EstadosRegistroHabito','Unidades','TipoImpactoDiversion','AmbitoAfectado','CapacidadAfectada','VistaDashboard','EscenarioSemana','Activo'])
    data={'A':['recurrente','bloque_simple'],'B':['fijo_inevitable','recurrente_configurable','seguimiento_sin_tiempo'],'C':['SI','NO'],'D':['SI','NO'],'E':['diaria','semanal','mensual','personalizada'],'F':['alta','media','baja'],'G':['critica','alta','media','baja'],'H':['no_iniciado','en_progreso','pausado','completado','cancelado'],'I':['no_iniciada','en_progreso','bloqueada','completada'],'J':['completado','parcial','omitido','no_aplica'],'K':['horas','sesiones','dias_exitosos','veces','minutos'],'L':['bloqueo_total','bloqueo_horas','reduccion_parcial'],'M':['todo_tiempo','solo_trabajo','solo_proyectos','categorias_especificas'],'N':['total_real','solo_proyectos'],'O':['semanal','mensual','anual'],'P':['base','vacaciones','pico','bloqueo_parcial'],'Q':['SI','NO']}
    for c,vals in data.items():
        for r,v in enumerate(vals,2): ws[f'{c}{r}']=v
    ws.freeze_panes='A2'; widths(ws,{'A':18,'B':24,'C':10,'D':14,'E':16,'F':12,'G':14,'H':22,'I':18,'J':24,'K':14,'L':22,'M':24,'N':18,'O':16,'P':18,'Q':10})

def act(wb):
    ws=wb.create_sheet('Actividades'); h=['ID','Nombre','Tipo','TipoBloque','EsFijo','ConsumeTiempo','Categoria','Subcategoria','Frecuencia','DiasSemana','VecesPorSemana','DuracionMin','TrasladoMinIda','TrasladoMinVuelta','DuracionTotalHorasCalc','MetaPeriodo','Unidad','Prioridad','PrioridadReal','Inicio','Fin','Activo','CapacidadAfectada','EscenarioSemana','ColorTag','Notas','Cumplimiento30d']; hdr(ws,1,h)
    ws.append(['ACT-001','Comidas e higiene','recurrente','fijo_inevitable','SI','SI','Personal','Rutina','diaria','1,2,3,4,5,6,7',7,120,0,0,'',1,'horas','alta','alta','2026-01-01','','SI','total_real','base','',''])
    ws.append(['ACT-002','Ejercicio presencial','recurrente','recurrente_configurable','NO','SI','Salud','Ejercicio','semanal','1,3,5',3,60,20,20,'',3,'sesiones','media','media','2026-01-01','','SI','solo_proyectos','base','',''])
    ws.append(['ACT-003','Lectura','recurrente','seguimiento_sin_tiempo','NO','NO','Desarrollo','Conocimiento','diaria','1,2,3,4,5,6,7',7,30,0,0,'',1,'sesiones','media','baja','2026-01-01','','SI','solo_proyectos','base','',''])
    for r in range(2,501):
        ws[f'O{r}']=f'=IF($A{r}="","",($L{r}+$M{r}+$N{r})/60)'
        ws[f'AA{r}']=f'=IF($A{r}="","",IF(COUNTIFS(RegistroHabitos!$C:$C,$A{r},RegistroHabitos!$B:$B,">="&TODAY()-30)=0,"",COUNTIFS(RegistroHabitos!$C:$C,$A{r},RegistroHabitos!$B:$B,">="&TODAY()-30,RegistroHabitos!$D:$D,"completado")/COUNTIFS(RegistroHabitos!$C:$C,$A{r},RegistroHabitos!$B:$B,">="&TODAY()-30)))'
    ws['T2'].number_format='yyyy-mm-dd'; ws.freeze_panes='A2'; widths(ws,{'A':11,'B':24,'C':14,'D':22,'E':9,'F':14,'G':14,'H':16,'I':14,'J':15,'K':14,'L':12,'M':14,'N':16,'O':18,'P':12,'Q':12,'R':11,'S':13,'T':12,'U':12,'V':9,'W':17,'X':16,'Y':10,'Z':28,'AA':16})

def pro(wb):
    ws=wb.create_sheet('Proyectos'); h=['ProyectoID','Nombre','Categoria','Subcategoria','Estado','Prioridad','PrioridadReal','Inicio','FechaObjetivo','HorasTotalesEstimadas','HorasAsignadasSemana','TieneMantenimiento','HorasMantenimientoSemana','DependenciaProyectoID','Avance%','Activo','CapacidadAfectada','EscenarioSemana','Notas','HorasRestantesCalc','SemanasRestantesCalc','FechaProbableFinCalc','AlertaCarga','ScoreDashboard']; hdr(ws,1,h)
    ws.append(['PRJ-001','MVP App Personal','Negocio','Producto','en_progreso','alta','alta','2026-01-01','',120,6,'NO',0,'',0.2,'SI','solo_proyectos','base',''])
    for r in range(2,501):
        ws[f'T{r}']=f'=IF($A{r}="","",MAX(0,$J{r}*(1-$O{r})))'; ws[f'U{r}']=f'=IF($A{r}="","",IF($K{r}>0,$T{r}/$K{r},""))'; ws[f'V{r}']=f'=IF($A{r}="","",IF(OR($H{r}="",$U{r}=""),"",$H{r}+ROUNDUP($U{r}*7,0)))'; ws[f'W{r}']=f'=IF($A{r}="","",IF($K{r}=0,"Sin asignacion",IF($K{r}>30,"Riesgo alta carga","OK")))'; ws[f'X{r}']=f'=IF($A{r}="","",IF(AND($P{r}="SI",$E{r}<>"completado"),((IF($G{r}="critica",4,IF($G{r}="alta",3,IF($G{r}="media",2,1)))*(1-$O{r}))+ROW()/100000),0))'
    ws['H2'].number_format='yyyy-mm-dd'; ws.freeze_panes='A2'; widths(ws,{'A':13,'B':24,'C':14,'D':15,'E':14,'F':10,'G':13,'H':12,'I':14,'J':18,'K':19,'L':18,'M':20,'N':19,'O':9,'P':9,'Q':17,'R':16,'S':26,'T':17,'U':20,'V':20,'W':18,'X':15})

def fas(wb):
    ws=wb.create_sheet('FasesProyecto'); h=['FaseID','ProyectoID','Orden','NombreFase','Estado','HorasEstimadas','HorasAsignadasSemana','DependenciaFaseID','InicioTentativo','FinProbableCalculado','Avance%','Notas','HorasRestantesCalc','SemanasRestantesCalc']; hdr(ws,1,h)
    ws.append(['FAS-001','PRJ-001',1,'Analisis','en_progreso',20,5,'','2026-01-01','',0.2,''])
    for r in range(2,501): ws[f'J{r}']=f'=IF($A{r}="","",IF(OR($I{r}="",$G{r}<=0,$F{r}=""),"",$I{r}+ROUNDUP(($F{r}*(1-$K{r}))/$G{r}*7,0)))'; ws[f'M{r}']=f'=IF($A{r}="","",MAX(0,$F{r}*(1-$K{r})))'; ws[f'N{r}']=f'=IF($A{r}="","",IF($G{r}>0,$M{r}/$G{r},""))'
    ws['I2'].number_format='yyyy-mm-dd'; ws.freeze_panes='A2'; widths(ws,{'A':12,'B':13,'C':8,'D':22,'E':16,'F':15,'G':20,'H':18,'I':14,'J':20,'K':10,'L':24,'M':18,'N':18})

def eve(wb):
    ws=wb.create_sheet('EventosDiversion'); h=['EventoID','Nombre','TipoImpacto','Inicio','Fin','HorasBloqueadas','PorcentajeReduccion','AmbitoAfectado','CapacidadAfectada','EscenarioSemana','Activo','Notas']; hdr(ws,1,h)
    ws.append(['DIV-001','Vacaciones familiares','bloqueo_total','2026-07-01','2026-07-07',0,0,'todo_tiempo','solo_proyectos','vacaciones','SI',''])
    ws.append(['DIV-002','Concierto','bloqueo_horas','2026-04-22','2026-04-22',5,0,'solo_proyectos','solo_proyectos','base','SI',''])
    ws.freeze_panes='A2'; widths(ws,{'A':12,'B':24,'C':20,'D':14,'E':14,'F':16,'G':20,'H':22,'I':18,'J':16,'K':10,'L':30})

def reg(wb):
    ws=wb.create_sheet('RegistroHabitos'); hdr(ws,1,['RegistroID','Fecha','ActividadID','Estado','Valor','Comentarios'])
    ws.append(['REG-001','2026-01-02','ACT-003','completado',1,''])
    ws.freeze_panes='A2'; widths(ws,{'A':12,'B':14,'C':14,'D':16,'E':10,'F':40})

def cal(wb):
    ws=wb.create_sheet('Calendario')
    h=['Fecha','Dia','DiaNum','SemanaISO','Mes','Año','HorasBase','HorasSueno','HorasBloquesFijosInevitables','HorasRecurrentesConfigurables','HorasBloqueadasExtraordinarias','CapacidadNetaDia','HorasProyectosDia','SaldoLibreRealDia','UtilizacionProyectos','Saturacion','ClaveSemana','ClaveMes','CapacidadTotalDia','HorasComprometidasTotales']
    hdr(ws,1,h)
    maxr=3661
    for r in range(2,maxr+1):
        ws[f'A{r}']='=IF(ROW()-2<Config!$B$3,Config!$B$2+ROW()-2,"")'
        ws[f'B{r}']=f'=IF($A{r}="","",CHOOSE(WEEKDAY($A{r},2),"Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo"))'
        ws[f'C{r}']=f'=IF($A{r}="","",WEEKDAY($A{r},2))'; ws[f'D{r}']=f'=IF($A{r}="","",ISOWEEKNUM($A{r}))'; ws[f'E{r}']=f'=IF($A{r}="","",MONTH($A{r}))'; ws[f'F{r}']=f'=IF($A{r}="","",YEAR($A{r}))'
        ws[f'G{r}']=f'=IF($A{r}="","",Config!$B$4)'; ws[f'H{r}']=f'=IF($A{r}="","",Config!$B$5)'
        scenA='IF(Config!$B$12="base",((Actividades!$X$2:$X$500="base")+(Actividades!$X$2:$X$500="")),((Actividades!$X$2:$X$500="base")+(Actividades!$X$2:$X$500="")+(Actividades!$X$2:$X$500=Config!$B$12)))'
        freq=f'((Actividades!$I$2:$I$500="diaria")+((Actividades!$I$2:$I$500="semanal")*((Actividades!$J$2:$J$500<>"")*ISNUMBER(SEARCH(","&$C{r}&",",","&Actividades!$J$2:$J$500&","))+(Actividades!$J$2:$J$500="")*(Actividades!$K$2:$K$500/7)))+((Actividades!$I$2:$I$500="personalizada")*ISNUMBER(SEARCH(","&$C{r}&",",","&Actividades!$J$2:$J$500&",")))+((Actividades!$I$2:$I$500="mensual")*(DAY($A{r})=DAY(Actividades!$T$2:$T$500))))'
        ws[f'I{r}']=f'=IF($A{r}="","",Config!$B$6+SUMPRODUCT((Actividades!$V$2:$V$500="SI")*(Actividades!$F$2:$F$500="SI")*(Actividades!$D$2:$D$500="fijo_inevitable")*(Actividades!$T$2:$T$500<=$A{r})*((Actividades!$U$2:$U$500="")+(Actividades!$U$2:$U$500>=$A{r}))*({scenA})*({freq})*Actividades!$O$2:$O$500))'
        ws[f'J{r}']=f'=IF($A{r}="","",SUMPRODUCT((Actividades!$V$2:$V$500="SI")*(Actividades!$F$2:$F$500="SI")*(Actividades!$D$2:$D$500="recurrente_configurable")*(Actividades!$T$2:$T$500<=$A{r})*((Actividades!$U$2:$U$500="")+(Actividades!$U$2:$U$500>=$A{r}))*({scenA})*({freq})*Actividades!$O$2:$O$500))'
        scenE='IF(Config!$B$12="base",((EventosDiversion!$J$2:$J$500="base")+(EventosDiversion!$J$2:$J$500="")),((EventosDiversion!$J$2:$J$500="base")+(EventosDiversion!$J$2:$J$500="")+(EventosDiversion!$J$2:$J$500=Config!$B$12)))'
        base=f'MAX(0,$G{r}-$H{r}-$I{r}-$J{r})'
        ws[f'K{r}']=f'=IF($A{r}="","",IF(SUMPRODUCT((EventosDiversion!$K$2:$K$500="SI")*(EventosDiversion!$C$2:$C$500="bloqueo_total")*(EventosDiversion!$D$2:$D$500<=$A{r})*(EventosDiversion!$E$2:$E$500>=$A{r})*(EventosDiversion!$H$2:$H$500<>"categorias_especificas")*({scenE}))>0,{base},MIN({base},SUMPRODUCT((EventosDiversion!$K$2:$K$500="SI")*(EventosDiversion!$C$2:$C$500="bloqueo_horas")*(EventosDiversion!$D$2:$D$500<=$A{r})*(EventosDiversion!$E$2:$E$500>=$A{r})*(EventosDiversion!$H$2:$H$500<>"categorias_especificas")*({scenE})*EventosDiversion!$F$2:$F$500)+{base}*MIN(1,SUMPRODUCT((EventosDiversion!$K$2:$K$500="SI")*(EventosDiversion!$C$2:$C$500="reduccion_parcial")*(EventosDiversion!$D$2:$D$500<=$A{r})*(EventosDiversion!$E$2:$E$500>=$A{r})*(EventosDiversion!$H$2:$H$500<>"categorias_especificas")*({scenE})*EventosDiversion!$G$2:$G$500)))))'
        scenP='IF(Config!$B$12="base",((Proyectos!$R$2:$R$500="base")+(Proyectos!$R$2:$R$500="")),((Proyectos!$R$2:$R$500="base")+(Proyectos!$R$2:$R$500="")+(Proyectos!$R$2:$R$500=Config!$B$12)))'
        ws[f'L{r}']=f'=IF($A{r}="","",MAX(0,$G{r}-$H{r}-$I{r}-$J{r}-$K{r}))'; ws[f'M{r}']=f'=IF($A{r}="","",SUMPRODUCT((Proyectos!$P$2:$P$500="SI")*(Proyectos!$H$2:$H$500<=$A{r})*({scenP})*(((Proyectos!$E$2:$E$500<>"completado")*(Proyectos!$K$2:$K$500/7))+((Proyectos!$L$2:$L$500="SI")*(Proyectos!$M$2:$M$500/7)))))'
        ws[f'N{r}']=f'=IF($A{r}="","",$L{r}-$M{r})'; ws[f'O{r}']=f'=IF($A{r}="","",IF($L{r}=0,IF($M{r}>0,1,0),$M{r}/$L{r}))'; ws[f'P{r}']=f'=IF($A{r}="","",IF($N{r}<0,"SATURADO",IF($N{r}<Config!$B$11,"RIESGO","OK")))'; ws[f'Q{r}']=f'=IF($A{r}="","",TEXT($F{r},"0000")&"-W"&TEXT($D{r},"00"))'; ws[f'R{r}']=f'=IF($A{r}="","",TEXT($A{r},"yyyy-mm"))'; ws[f'S{r}']=f'=IF($A{r}="","",MAX(0,$G{r}-$H{r}))'; ws[f'T{r}']=f'=IF($A{r}="","",$I{r}+$J{r}+$K{r}+$M{r})'
    ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:T{maxr}'; widths(ws,{'A':12,'B':12,'C':9,'D':11,'E':8,'F':8,'G':11,'H':11,'I':22,'J':22,'K':22,'L':16,'M':16,'N':16,'O':14,'P':12,'Q':13,'R':11,'S':16,'T':22})
    for r in range(2,maxr+1): ws[f'A{r}'].number_format='yyyy-mm-dd'; ws[f'O{r}'].number_format='0%'
    ws.conditional_formatting.add(f'N2:N{maxr}',CellIsRule(operator='lessThan',formula=['0'],fill=PatternFill('solid',fgColor='F8CBAD'))); ws.conditional_formatting.add(f'N2:N{maxr}',CellIsRule(operator='between',formula=['0','1.999'],fill=PatternFill('solid',fgColor='FFE699'))); ws.conditional_formatting.add(f'N2:N{maxr}',CellIsRule(operator='greaterThanOrEqual',formula=['2'],fill=PatternFill('solid',fgColor='C6EFCE')))

def dash(wb):
    ws=wb.create_sheet('Dashboard'); ws['A1']='Ninja Tracker MVP - Calculadora de Tiempo Disponible Real'; ws['A1'].font=Font(size=15,bold=True,color='1F4E78')
    ws['A3']='Vista'; ws['B3']='semanal'; ws['A4']='FechaReferencia'; ws['B4']='=TODAY()'; ws['A5']='EscenarioSemana'; ws['B5']='base'; ws['B4'].number_format='yyyy-mm-dd'
    for c in ['A3','A4','A5']: ws[c].fill=SF; ws[c].font=Font(bold=True)
    labels=['Capacidad total periodo (h)','Capacidad neta periodo (h)','Comprometido en proyectos (h)','Saldo libre real (h)','Maximo asignable sin sobrecarga (h)','Dias saturados','Utilizacion real']
    for i,t in enumerate(labels,7): ws[f'A{i}']=t; ws[f'A{i}'].fill=SF; ws[f'A{i}'].font=Font(bold=True)
    ws['B7']=ps('S'); ws['B8']=ps('L'); ws['B9']=ps('M'); ws['B10']='=B8-B9'; ws['B11']='=MAX(0,B10)'; ws['B12']='=IF($B$3="semanal",COUNTIFS(Calendario!$D:$D,ISOWEEKNUM($B$4),Calendario!$F:$F,YEAR($B$4),Calendario!$N:$N,"<0"),IF($B$3="mensual",COUNTIFS(Calendario!$E:$E,MONTH($B$4),Calendario!$F:$F,YEAR($B$4),Calendario!$N:$N,"<0"),COUNTIFS(Calendario!$F:$F,YEAR($B$4),Calendario!$N:$N,"<0")))'; ws['B13']='=IF(B8=0,0,B9/B8)'; ws['B13'].number_format='0%'
    for r in [7,8,9,10,11,12]: ws[f'B{r}'].number_format='0.0'
    ws['D7']='Estado de carga'; ws['E7']='=IF(B10<0,"SATURADO",IF(B10<Config!$B$11*7,"RIESGO","ESTABLE"))'; ws['E7'].font=Font(bold=True); ws['D8']='Escenario activo'; ws['E8']='=B5'; ws['D10']='Bloques no-proyecto (h)'; ws['E10']='=B18+B19+B20'; ws['D11']='Compromiso total (h)'; ws['E11']='=B22'
    ws.conditional_formatting.add('E7',CellIsRule(operator='equal',formula=['"SATURADO"'],fill=PatternFill('solid',fgColor='F8CBAD'))); ws.conditional_formatting.add('E7',CellIsRule(operator='equal',formula=['"RIESGO"'],fill=PatternFill('solid',fgColor='FFE699'))); ws.conditional_formatting.add('E7',CellIsRule(operator='equal',formula=['"ESTABLE"'],fill=PatternFill('solid',fgColor='C6EFCE')))
    ws['A16']='Consumo por tipo de bloque (periodo activo)'; ws['A16'].font=Font(bold=True,color='1F4E78'); hdr(ws,17,['Bloque','Horas','% de capacidad neta']); ws['A18']='Fijos inevitables'; ws['A19']='Recurrentes configurables'; ws['A20']='Eventos extraordinarios'; ws['A21']='Proyectos'; ws['A22']='Total comprometido'
    ws['B18']=ps('I'); ws['B19']=ps('J'); ws['B20']=ps('K'); ws['B21']=ps('M'); ws['B22']='=SUM(B18:B21)'
    for r in range(18,23): ws[f'C{r}']=f'=IF($B$8=0,0,B{r}/$B$8)'; ws[f'B{r}'].number_format='0.0'; ws[f'C{r}'].number_format='0%'
    ws['A25']='Top 5 proyectos activos'; ws['A25'].font=Font(bold=True,color='1F4E78'); hdr(ws,26,['Proyecto','Estado','Avance','Fin probable','_score'])
    for r in range(27,32): ws[f'E{r}']=f'=IFERROR(LARGE(Proyectos!$X$2:$X$500,ROW()-26),"")'; ws[f'A{r}']=f'=IF($E{r}="","",INDEX(Proyectos!$B$2:$B$500,MATCH($E{r},Proyectos!$X$2:$X$500,0)))'; ws[f'B{r}']=f'=IF($E{r}="","",INDEX(Proyectos!$E$2:$E$500,MATCH($E{r},Proyectos!$X$2:$X$500,0)))'; ws[f'C{r}']=f'=IF($E{r}="","",INDEX(Proyectos!$O$2:$O$500,MATCH($E{r},Proyectos!$X$2:$X$500,0)))'; ws[f'D{r}']=f'=IF($E{r}="","",INDEX(Proyectos!$V$2:$V$500,MATCH($E{r},Proyectos!$X$2:$X$500,0)))'; ws[f'C{r}'].number_format='0%'; ws[f'D{r}'].number_format='yyyy-mm-dd'
    ws['A34']='Habitos con mejor cumplimiento (30 dias)'; ws['A34'].font=Font(bold=True,color='1F4E78'); hdr(ws,35,['Habito','Cumplimiento','_score'])
    for r in range(36,41): ws[f'C{r}']=f'=IFERROR(LARGE(Actividades!$AA$2:$AA$500,ROW()-35),"")'; ws[f'A{r}']=f'=IF($C{r}="","",INDEX(Actividades!$B$2:$B$500,MATCH($C{r},Actividades!$AA$2:$AA$500,0)))'; ws[f'B{r}']=f'=$C{r}'; ws[f'B{r}'].number_format='0%'
    ws['E34']='Habitos rezagados (30 dias)'; ws['E34'].font=Font(bold=True,color='1F4E78'); ws['E35']='Habito'; ws['F35']='Cumplimiento'; ws['G35']='_score'
    for c in ['E35','F35','G35']: ws[c].fill=HF; ws[c].font=HFF; ws[c].alignment=Alignment(horizontal='center',vertical='center')
    for r in range(36,41): ws[f'G{r}']=f'=IFERROR(SMALL(Actividades!$AA$2:$AA$500,ROW()-35),"")'; ws[f'E{r}']=f'=IF($G{r}="","",INDEX(Actividades!$B$2:$B$500,MATCH($G{r},Actividades!$AA$2:$AA$500,0)))'; ws[f'F{r}']=f'=$G{r}'; ws[f'F{r}'].number_format='0%'
    ws.column_dimensions['C'].hidden=True; ws.column_dimensions['E'].hidden=True; ws.column_dimensions['G'].hidden=True; ws.freeze_panes='A7'; widths(ws,{'A':34,'B':20,'C':12,'D':22,'E':24,'F':18,'G':12})

def imp(wb):
    ws=wb.create_sheet('ImportCSV_Map'); hdr(ws,1,['Entidad','ColumnaCSV','Obligatoria','Formato','Ejemplo','ReglaValidacion','Notas'])
    rows=[('Actividades','ID','SI','texto','ACT-010','unico','Identificador'),('Actividades','TipoBloque','SI','catalogo','fijo_inevitable','Catálogos!B','Clasifica fijo/recurrente'),('Actividades','EsFijo','SI','SI/NO','SI','Catálogos!C','Bloque inevitable'),('Actividades','ConsumeTiempo','SI','SI/NO','SI','Catálogos!D','Descuenta capacidad'),('Actividades','DuracionMin','SI','numero','60','>=0','Duracion base'),('Actividades','TrasladoMinIda','OPC','numero','20','>=0','Traslado explicito'),('Actividades','TrasladoMinVuelta','OPC','numero','20','>=0','Traslado explicito'),('Actividades','VecesPorSemana','OPC','numero','3','>=0','Si no hay DiasSemana'),('Actividades','CapacidadAfectada','SI','catalogo','solo_proyectos','Catálogos!N','Alcance de capacidad'),('Actividades','EscenarioSemana','OPC','catalogo','base','Catálogos!P','Base o escenario'),('Proyectos','ProyectoID','SI','texto','PRJ-010','unico','ID'),('Proyectos','PrioridadReal','SI','catalogo','critica','Catálogos!G','Prioridad de capacidad'),('Proyectos','HorasTotalesEstimadas','SI','numero','80','>0','Estimado'),('Proyectos','HorasAsignadasSemana','SI','numero','6','>=0','Carga semanal'),('Proyectos','EscenarioSemana','OPC','catalogo','base','Catálogos!P','Escenario'),('FasesProyecto','FaseID','SI','texto','FAS-010','unico',''),('FasesProyecto','ProyectoID','SI','texto','PRJ-010','debe existir',''),('EventosDiversion','EventoID','SI','texto','DIV-010','unico',''),('EventosDiversion','TipoImpacto','SI','catalogo','bloqueo_horas','Catálogos!L','Tipo de bloqueo'),('EventosDiversion','EscenarioSemana','OPC','catalogo','vacaciones','Catálogos!P','Simulacion semanal'),('EventosDiversion','Activo','SI','SI/NO','SI','Catálogos!Q','Participa en calculo')]
    for i,row in enumerate(rows,2):
        for j,v in enumerate(row,1): ws.cell(i,j,v)
    ws.freeze_panes='A2'; widths(ws,{'A':20,'B':24,'C':12,'D':12,'E':20,'F':28,'G':38}); grid(ws,2,1+len(rows),7)

def val(wb):
    c="'Catálogos'"
    ws=wb['Actividades']; arr=[(f'={c}!$A$2:$A$20','C2:C500'),(f'={c}!$B$2:$B$20','D2:D500'),(f'={c}!$C$2:$C$20','E2:E500'),(f'={c}!$D$2:$D$20','F2:F500'),(f'={c}!$E$2:$E$20','I2:I500'),(f'={c}!$K$2:$K$20','Q2:Q500'),(f'={c}!$F$2:$F$20','R2:R500'),(f'={c}!$G$2:$G$20','S2:S500'),(f'={c}!$Q$2:$Q$20','V2:V500'),(f'={c}!$N$2:$N$20','W2:W500'),(f'={c}!$P$2:$P$20','X2:X500')]
    for f,r in arr: d=DataValidation(type='list',formula1=f,allow_blank=True); ws.add_data_validation(d); d.add(r)
    ws=wb['Proyectos']; arr=[(f'={c}!$H$2:$H$20','E2:E500'),(f'={c}!$F$2:$F$20','F2:F500'),(f'={c}!$G$2:$G$20','G2:G500'),(f'={c}!$Q$2:$Q$20','L2:L500'),(f'={c}!$Q$2:$Q$20','P2:P500'),(f'={c}!$N$2:$N$20','Q2:Q500'),(f'={c}!$P$2:$P$20','R2:R500')]
    for f,r in arr: d=DataValidation(type='list',formula1=f,allow_blank=True); ws.add_data_validation(d); d.add(r)
    ws=wb['FasesProyecto']; d=DataValidation(type='list',formula1=f'={c}!$I$2:$I$20',allow_blank=True); ws.add_data_validation(d); d.add('E2:E500')
    ws=wb['EventosDiversion']; arr=[(f'={c}!$L$2:$L$20','C2:C500'),(f'={c}!$M$2:$M$20','H2:H500'),(f'={c}!$N$2:$N$20','I2:I500'),(f'={c}!$P$2:$P$20','J2:J500'),(f'={c}!$Q$2:$Q$20','K2:K500')]
    for f,r in arr: d=DataValidation(type='list',formula1=f,allow_blank=True); ws.add_data_validation(d); d.add(r)
    ws=wb['RegistroHabitos']; d=DataValidation(type='list',formula1=f'={c}!$J$2:$J$20',allow_blank=True); ws.add_data_validation(d); d.add('D2:D500')
    ws=wb['Dashboard']; d=DataValidation(type='list',formula1=f'={c}!$O$2:$O$20',allow_blank=False); ws.add_data_validation(d); d.add('B3'); d2=DataValidation(type='list',formula1=f'={c}!$P$2:$P$20',allow_blank=False); ws.add_data_validation(d2); d2.add('B5')

def post(wb):
    for s,c in {'Actividades':27,'Proyectos':24,'FasesProyecto':14,'EventosDiversion':12,'RegistroHabitos':6}.items():
        ws=wb[s]; grid(ws,2,500,c); ws.auto_filter.ref=ws.dimensions

def formulario(wb):
    """Hoja de entrada guiada: reemplaza edición directa en Actividades/Proyectos/Fases."""
    ws = wb.create_sheet('Formulario_Entrada')
    ws['A1'] = 'FORMULARIO DE ENTRADA — Ninja Tracker'
    ws['A1'].font = Font(size=14, bold=True, color='1F4E78')

    # --- Sección Básicos / Recurrentes ---
    ws['A3'] = 'ACTIVIDADES (Básicos y Recurrentes)'
    ws['A3'].font = Font(bold=True, color='FFFFFF')
    ws['A3'].fill = PatternFill('solid', fgColor='1F4E78')
    h_act = ['ID','Nombre','Tipo','Escenarios','DiaSemana','HoraInicio','HoraFin','HorasPorSemana_calc','Activo','Notas']
    hdr(ws, 4, h_act)
    # Fórmula HorasPorSemana desde franjas: (HoraFin - HoraInicio) sumada por ID
    # En este formulario simplificado una fila = una franja
    for r in range(5, 105):
        ws[f'H{r}'] = f'=IF(OR($E{r}="", $F{r}="", $G{r}=""), "", ($G{r}-$F{r})*24)'
        ws[f'H{r}'].number_format = '0.00'

    ws['A3'].fill = PatternFill('solid', fgColor='1F4E78')
    ws.merge_cells('A3:J3')

    # Ejemplos
    ws.append([]); ws.append([])  # skip to row 5
    # Already at row 5 via header at 4 — insert sample data directly
    ws['A5'] = 'ACT-001'; ws['B5'] = 'Sueño'; ws['C5'] = 'basico'
    ws['D5'] = 'base,vacaciones,pico,bloqueo_parcial'; ws['E5'] = 'lunes'
    ws['F5'] = 0.0; ws['G5'] = 0.333  # 00:00 - 08:00 como fracción de día
    ws['F5'].number_format = '[h]:mm'; ws['G5'].number_format = '[h]:mm'
    ws['I5'] = 'SI'

    # --- Sección Proyectos ---
    ws['A108'] = 'PROYECTOS'
    ws['A108'].font = Font(bold=True, color='FFFFFF')
    ws['A108'].fill = PatternFill('solid', fgColor='1F4E78')
    ws.merge_cells('A108:M108')
    h_pro = ['ID','Nombre','Estado','HorasTotales','Avance%','Inicio','FinFijo','DependeDe','DiaSemana','HoraInicio','HoraFin','HorasPorSemana_calc','FinEstimado_calc']
    hdr(ws, 109, h_pro)
    for r in range(110, 160):
        # HorasPorSemana_calc: suma de franjas — simplificado aquí como (HoraFin-HoraInicio)*24
        ws[f'L{r}'] = f'=IF(OR($I{r}="", $J{r}="", $K{r}=""), "", ($K{r}-$J{r})*24)'
        ws[f'L{r}'].number_format = '0.00'
        # FinEstimado: Inicio + HorasRestantes/HorasPorSemana/5*7
        ws[f'M{r}'] = (
            f'=IF(OR($F{r}="", $L{r}="", $L{r}=0, $D{r}=""), "", '
            f'$F{r}+ROUNDUP(($D{r}*(1-IF($E{r}="",0,$E{r})))/$L{r}*7, 0))'
        )
        ws[f'M{r}'].number_format = 'yyyy-mm-dd'

    # --- Sección Fases ---
    ws['A163'] = 'FASES DE PROYECTO'
    ws['A163'].font = Font(bold=True, color='FFFFFF')
    ws['A163'].fill = PatternFill('solid', fgColor='1F4E78')
    ws.merge_cells('A163:L163')
    h_fas = ['FaseID','ProyectoID','Orden','Nombre','HorasEstimadas','Avance%','DependeDeFaseID','DiaSemana','HoraInicio','HoraFin','HorasPorSemana_calc','InicioCalculado','FinCalculado']
    hdr(ws, 164, h_fas)
    for r in range(165, 215):
        ws[f'K{r}'] = f'=IF(OR($H{r}="", $I{r}="", $J{r}=""), "", ($J{r}-$I{r})*24)'
        ws[f'K{r}'].number_format = '0.00'
        # FinCalculado: InicioCalculado + HorasEstimadas*(1-avance)/HorasPorSemana*7
        ws[f'M{r}'] = (
            f'=IF(OR($L{r}="", $K{r}="", $K{r}=0, $E{r}=""), "", '
            f'$L{r}+ROUNDUP(($E{r}*(1-IF($F{r}="",0,$F{r})))/$K{r}*7, 0))'
        )
        ws[f'M{r}'].number_format = 'yyyy-mm-dd'

    # Nota instruccional
    ws['A218'] = 'INSTRUCCIONES'
    ws['A218'].font = Font(bold=True)
    instrucciones = [
        ('1.', 'Llena los bloques Básicos/Recurrentes con tus franjas (una fila por franja por día).'),
        ('2.', 'Llena los Proyectos con horas totales, avance y al menos una franja horaria.'),
        ('3.', 'Llena las Fases. El campo InicioCalculado se actualiza automáticamente (Apps Script).'),
        ('4.', 'Usa el menú "Ninja Tracker > Recalcular dependencias" para actualizar las fechas en cascada.'),
        ('5.', 'Las hojas Actividades, Proyectos y FasesProyecto son el motor de cálculo — no editarlas directamente.'),
    ]
    for i, (num, txt) in enumerate(instrucciones, 219):
        ws[f'A{i}'] = num; ws[f'B{i}'] = txt; ws[f'A{i}'].font = Font(bold=True)

    widths(ws, {'A': 13, 'B': 28, 'C': 16, 'D': 30, 'E': 14, 'F': 12, 'G': 12, 'H': 18,
                'I': 14, 'J': 12, 'K': 12, 'L': 18, 'M': 18})
    ws.freeze_panes = 'A5'

if __name__=='__main__':
    wb=Workbook(); cfg(wb.active); cat(wb); act(wb); pro(wb); fas(wb); eve(wb); reg(wb); cal(wb); dash(wb); imp(wb); val(wb); formulario(wb); post(wb)
    OUT.parent.mkdir(parents=True,exist_ok=True); wb.save(OUT); print(f'Archivo generado: {OUT}')
